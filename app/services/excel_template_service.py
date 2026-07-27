"""Service for generating Excel templates for waybill data entry."""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class ExcelTemplateService:
    """Generates Excel workbook templates for waybill bulk entry."""

    @staticmethod
    def generate_waybill_template() -> bytes:
        """Generate a waybill entry Excel template with headers, sample data, and styling.

        Returns:
            Raw bytes of the .xlsx file ready for download.
        """
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = "Waybills"
        ws.sheet_view.rightToLeft = True

        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Define columns
        columns = [
            ("نام فرستنده", "sender_name"),
            ("کد ملی فرستنده", "sender_national_code"),
            ("موبایل فرستنده", "sender_phone"),
            ("آدرس فرستنده", "sender_address"),
            ("نام گیرنده", "receiver_name"),
            ("کد ملی گیرنده", "receiver_national_code"),
            ("موبایل گیرنده", "receiver_phone"),
            ("آدرس گیرنده", "receiver_address"),
            ("استان مبدأ", "origin_province"),
            ("شهر مبدأ", "origin_city"),
            ("منطقه مبدأ", "origin_district"),
            ("آدرس مبدأ", "origin_address"),
            ("استان مقصد", "destination_province"),
            ("شهر مقصد", "destination_city"),
            ("منطقه مقصد", "destination_district"),
            ("آدرس مقصد", "destination_address"),
            ("نوع کالا", "cargo_type"),
            ("وزن بار (تن)", "cargo_weight"),
            ("تعداد بار", "cargo_count"),
            ("توضیحات کالا", "cargo_description"),
            ("کد ملی راننده", "driver_national_code"),
            ("تلفن راننده", "driver_phone"),
            ("پلاک ملی: دو رقم اول پلاک", "plate_first_two"),
            ("پلاک ملی: حرف پلاک", "plate_letter"),
            ("پلاک ملی: سه رقم پلاک", "plate_three"),
            ("پلاک ملی: دو رقم آخر پلاک", "plate_last_two"),
            ("هزینه حمل", "cost"),
            ("روش پرداخت", "payment_method"),
            ("نام کاربری اکانت ثبت", "account_username"),
            ("رمز عبور اکانت ثبت", "account_password"),
        ]

        # Write headers
        for col_idx, (header_name, _) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add sample data
        sample_data = [
            "علی احمدی",
            "1234567890",
            "09121234567",
            "تهران، خیابان ولیعصر",
            "مرضا رضایی",
            "0987654321",
            "09139876543",
            "اصفهان، خیابان چهارباغ",
            "تهران",
            "تهران",
            "مرکز",
            "تهران، میدان آزادی",
            "اصفهان",
            "اصفهان",
            "ناجوان",
            "اصفهان، خیابان آمادگاه",
            "مواد غذایی",
            "10.5",
            "5",
            "بار خشک",
            "1234567890",
            "09121111111",
            "11",
            "ع",
            "222",
            "33",
            "5000000",
            "نقدی",
            "user@example.com",
            "[رمز عبور اکانت - وارد کنید]",
        ]

        for col_idx, value in enumerate(sample_data, start=1):
            ws.cell(row=2, column=col_idx, value=value)

        # Set column widths
        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()
