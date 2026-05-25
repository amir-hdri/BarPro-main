"""API routes for manual waybill entry and Excel file upload."""

import logging

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form

from app.core.security import require_sensitive_auth
from app.schemas.waybill import OperationMode, WaybillMapRequest
from app.services.waybill_entry_service import excel_waybill_service, manual_waybill_service

router = APIRouter(prefix="/waybill", tags=["waybill-entry"])
logger = logging.getLogger(__name__)


@router.post(
    "/validate-manual-entry",
    dependencies=[Depends(require_sensitive_auth)],
)
async def validate_manual_entry(request: WaybillMapRequest):
    """اعتبارسنجی ورودی دستی بارنامه قبل از ارسال."""
    validation = manual_waybill_service.validate_manual_entry(request)
    return {
        "valid": validation["valid"],
        "errors": validation["errors"],
        "warnings": validation["warnings"],
        "field_count": validation["field_count"],
        "completed_fields": validation["completed_fields"],
        "completion_percent": round(
            validation["completed_fields"] / validation["field_count"] * 100, 1
        ) if validation["field_count"] > 0 else 0,
    }


@router.post(
    "/submit-manual-waybill",
    dependencies=[Depends(require_sensitive_auth)],
)
async def submit_manual_waybill(request: WaybillMapRequest):
    """ارسال دستی بارنامه با اعتبارسنجی اولیه و قرارگیری در صف برای اجرا."""
    from app.queue.queue_manager import queue_manager

    # Validate first
    validation = manual_waybill_service.validate_manual_entry(request)
    if not validation["valid"]:
        raise HTTPException(
            status_code=422,
            detail={
                "error": "VALIDATION_FAILED",
                "message": "اطلاعات وارد شده نامعتبر است",
                "errors": validation["errors"],
                "warnings": validation["warnings"],
            },
        )

    # Enqueue via queue manager
    try:
        task = await queue_manager.enqueue_waybill(request)
        return {
            "success": True,
            "message": "بارنامه با موفقیت در صف ثبت قرار گرفت",
            "result": task.model_dump(),
        }
    except HTTPException as exc:
        raise exc
    except Exception as exc:
        logger.exception("manual_waybill_submit_failed")
        raise HTTPException(
            status_code=500,
            detail={
                "error": "SUBMISSION_FAILED",
                "message": "خطا در قرارگیری بارنامه در صف",
                "detail": str(exc),
            },
        )


@router.post(
    "/parse-excel",
    dependencies=[Depends(require_sensitive_auth)],
)
async def parse_excel_file(
    file: UploadFile = File(..., description="فایل اکسل حاوی اطلاعات بارنامه"),
    operation_mode: OperationMode = Form(default=OperationMode.SAFE),
):
    """پردازش فایل اکسل و نمایش اطلاعات بارنامه‌ها."""
    # Validate file type
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail={
                "error": "INVALID_FILE_TYPE",
                "message": "فایل باید با فرمت اکسل (xlsx/xls) باشد",
            },
        )

    # Validate file size (max 10MB)
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(
            status_code=400,
            detail={
                "error": "FILE_TOO_LARGE",
                "message": "حجم فایل نباید بیشتر از 10 مگابایت باشد",
            },
        )
    
    # Reset file pointer
    await file.seek(0)

    # Parse Excel
    result = await excel_waybill_service.parse_excel_file(file, operation_mode)
    
    if not result["success"]:
        raise HTTPException(
            status_code=400,
            detail={
                "error": "EXCEL_PARSE_FAILED",
                "message": result.get("error", "خطا در پردازش فایل اکسل"),
            },
        )

    return {
        "success": True,
        "file_name": result["file_name"],
        "total_rows": result["total_rows"],
        "valid_waybills": result["valid_waybills"],
        "errors": result["errors"],
        "waybills_preview": [
            {
                "row": item["row"],
                "sender": item["waybill"].sender.name,
                "receiver": item["waybill"].receiver.name,
                "origin": f"{item['waybill'].origin.city}, {item['waybill'].origin.province}",
                "destination": f"{item['waybill'].destination.city}, {item['waybill'].destination.province}",
                "cargo_weight": item["waybill"].cargo.weight,
                "validation": item["validation"],
            }
            for item in result["waybills"][:10]  # Show first 10 as preview
        ],
        "error_details": result.get("error_details", [])[:10],
    }


@router.post(
    "/submit-excel-waybills",
    dependencies=[Depends(require_sensitive_auth)],
)
async def submit_excel_waybills(
    file: UploadFile = File(..., description="فایل اکسل حاوی اطلاعات بارنامه"),
    operation_mode: OperationMode = Form(default=OperationMode.SAFE),
    skip_invalid: bool = Form(default=True, description="رد کردن موارد نامعتبر"),
):
    """پردازش و ارسال گروهی بارنامه‌ها از فایل اکسل."""
    # Validate file type
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail={
                "error": "INVALID_FILE_TYPE",
                "message": "فایل باید با فرمت اکسل (xlsx/xls) باشد",
            },
        )

    # Process Excel file
    result = await excel_waybill_service.process_excel_waybills(
        file, operation_mode, skip_invalid
    )

    if not result["success"]:
        raise HTTPException(
            status_code=400,
            detail={
                "error": "EXCEL_PROCESS_FAILED",
                "message": result.get("error", "خطا در پردازش فایل اکسل"),
            },
        )

    return {
        "success": True,
        "message": f"پردازش {result['total_processed']} بارنامه انجام شد",
        "file_name": result["file_name"],
        "total_processed": result["total_processed"],
        "success_count": result["success_count"],
        "error_count": result["error_count"],
        "results": result["results"],
    }


@router.post(
    "/queue-excel-waybills",
    dependencies=[Depends(require_sensitive_auth)],
)
async def queue_excel_waybills(
    file: UploadFile = File(..., description="فایل اکسل حاوی اطلاعات بارنامه"),
    operation_mode: OperationMode = Form(default=OperationMode.SAFE),
    skip_invalid: bool = Form(default=True, description="رد کردن موارد نامعتبر"),
):
    """پردازش و افزودن گروهی بارنامه‌ها به صف."""
    from app.queue.queue_manager import queue_manager

    # Validate file type
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail={
                "error": "INVALID_FILE_TYPE",
                "message": "فایل باید با فرمت اکسل (xlsx/xls) باشد",
            },
        )

    # Parse Excel first
    parse_result = await excel_waybill_service.parse_excel_file(file, operation_mode)
    
    if not parse_result["success"]:
        raise HTTPException(
            status_code=400,
            detail={
                "error": "EXCEL_PARSE_FAILED",
                "message": parse_result.get("error", "خطا در پردازش فایل اکسل"),
            },
        )

    # Queue each valid waybill
    queued = []
    errors = []
    
    for item in parse_result["waybills"]:
        try:
            waybill = item["waybill"]
            
            # Enqueue
            task = await queue_manager.enqueue_waybill(waybill)
            queued.append({
                "row": item["row"],
                "task_id": task.get("task_id"),
                "status": "queued",
            })
        except Exception as exc:
            errors.append({
                "row": item["row"],
                "error": str(exc),
            })

    return {
        "success": True,
        "message": f"{len(queued)} بارنامه به صف اضافه شد",
        "file_name": parse_result["file_name"],
        "total_parsed": parse_result["total_rows"],
        "queued_count": len(queued),
        "error_count": len(errors),
        "queued_tasks": queued,
        "errors": errors,
    }


@router.get("/excel-template", tags=["templates"])
async def download_excel_template():
    """دانلود قالب اکسل برای ورود اطلاعات بارنامه."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "Waybills"
    ws.sheet_properties.rightToLeft = True

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Define columns
    columns = [
        ("نام فرستنده", "sender_name"),
        ("کد ملی فرستنده", "sender_national_code"),
        ("موبایل فرستنده", "sender_phone"),
        ("آدرس فرستنده", "sender_address"),
        ("نام گیرنده", "receiver_name"),
        ("کد ملی گیرنده", "receiver_national_code"),
        ("موبایل گیرنده", "receiver_phone"),
        ("آدرس گیرنده", "receiver_address"),
        ("استان مبدأ", "origin_province"),
        ("شهر مبدأ", "origin_city"),
        ("منطقه مبدأ", "origin_district"),
        ("آدرس مبدأ", "origin_address"),
        ("استان مقصد", "destination_province"),
        ("شهر مقصد", "destination_city"),
        ("منطقه مقصد", "destination_district"),
        ("آدرس مقصد", "destination_address"),
        ("نوع کالا", "cargo_type"),
        ("وزن بار (تن)", "cargo_weight"),
        ("تعداد بار", "cargo_count"),
        ("توضیحات کالا", "cargo_description"),
        ("کد ملی راننده", "driver_national_code"),
        ("تلفن راننده", "driver_phone"),
        ("پلاک ملی: دو رقم اول پلاک", "plate_first_two"),
        ("پلاک ملی: حرف پلاک", "plate_letter"),
        ("پلاک ملی: سه رقم پلاک", "plate_three"),
        ("پلاک ملی: دو رقم آخر پلاک", "plate_last_two"),
        ("هزینه حمل", "cost"),
        ("روش پرداخت", "payment_method"),
        ("نام کاربری اکانت ثبت", "account_username"),
        ("رمز عبور اکانت ثبت", "account_password"),
    ]

    # Write headers
    for col_idx, (header_name, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Add sample data
    sample_data = [
        "علی احمدی",
        "1234567890",
        "09121234567",
        "تهران، خیابان ولیعصر",
        "مرضا رضایی",
        "0987654321",
        "09139876543",
        "اصفهان، خیابان چهارباغ",
        "تهران",
        "تهران",
        "مرکز",
        "تهران، میدان آزادی",
        "اصفهان",
        "اصفهان",
        "ناجوان",
        "اصفهان، خیابان آمادگاه",
        "مواد غذایی",
        "10.5",
        "5",
        "بار خشک",
        "1234567890",
        "09121111111",
        "11",
        "ع",
        "222",
        "33",
        "5000000",
        "نقدی",
        "user@example.com",
        "password123",
    ]

    for col_idx, value in enumerate(sample_data, start=1):
        ws.cell(row=2, column=col_idx, value=value)

    # Set column widths
    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=waybill_template.xlsx"},
    )
